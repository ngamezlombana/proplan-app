import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime, timedelta
import re

def get_month_name(date_obj):
    months = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    return f"{date_obj.month:02d} - {months[date_obj.month - 1].capitalize()}"

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def generate_excel_bytes(opsData, colorMap):
    wb = openpyxl.Workbook()
    
    # ---------------------------------------------------------
    # HOJA 1: PLANEADOR
    # ---------------------------------------------------------
    ws_plan = wb.active
    ws_plan.title = "PLANEADOR"
    
    if not opsData:
        return None
        
    min_date_str = "9999-12-31"
    max_date_str = "0000-01-01"
    for op in opsData:
        for task in op['tasks']:
            if task['start'] < min_date_str: min_date_str = task['start']
            if task['end'] > max_date_str: max_date_str = task['end']
            
    if min_date_str == "9999-12-31":
        min_date_str = datetime.now().strftime('%Y-%m-%d')
        max_date_str = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')
    else:
        max_date = datetime.strptime(max_date_str, '%Y-%m-%d') + timedelta(days=5)
        max_date_str = max_date.strftime('%Y-%m-%d')
        
    min_date = datetime.strptime(min_date_str, '%Y-%m-%d')
    max_date = datetime.strptime(max_date_str, '%Y-%m-%d')
    
    dates = []
    curr = min_date
    while curr <= max_date:
        dates.append(curr.strftime('%Y-%m-%d'))
        curr += timedelta(days=1)
        
    headers = ['OP', 'DESCRIPCION', 'CLIENTE', 'ZONA', 'EJE', 'TIPO', 'FECHA FIN']
    for d in dates:
        dt = datetime.strptime(d, '%Y-%m-%d')
        headers.append(dt.strftime('%d/%m'))
        
    ws_plan.append(headers)
    
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num in range(1, len(headers) + 1):
        cell = ws_plan.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws_plan.freeze_panes = "D2"
    
    row_idx = 2
    for op in opsData:
        d = op['details']
        ff = d.get('fecha_fin', '')
        if ff and ff != '9999-12-31':
            ff_parts = ff.split('-')
            ff_display = f"{ff_parts[2]}/{ff_parts[1]}/{ff_parts[0]}"
        else:
            ff_display = ""
            
        row_data = [
            d.get('op', ''),
            d.get('desc', ''),
            d.get('cliente', ''),
            d.get('zona', ''),
            d.get('eje', ''),
            d.get('tipo', ''),
            ff_display
        ]
        row_data.extend([""] * len(dates))
        ws_plan.append(row_data)
        
        occupied = {}
        for task in op['tasks']:
            color_hex = colorMap.get(task['group'], '#3b82f6').replace('#', '')
            machine_name = task['machine']
            if re.match(r'^0+$', machine_name):
                machine_name = '0'
            for date_str in task['all_dates']:
                occupied[date_str] = (machine_name, color_hex)
                
        for i, date_str in enumerate(dates):
            if date_str in occupied:
                machine, color_hex = occupied[date_str]
                col_num = 8 + i
                cell = ws_plan.cell(row=row_idx, column=col_num)
                cell.value = machine
                r = int(color_hex[0:2], 16)
                g = int(color_hex[2:4], 16)
                b = int(color_hex[4:6], 16)
                yiq = ((r*299)+(g*587)+(b*114))/1000
                font_color = "000000" if yiq >= 128 else "FFFFFF"
                cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                cell.font = Font(color=font_color, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        row_idx += 1
        
    ws_plan.column_dimensions['A'].width = 15
    ws_plan.column_dimensions['B'].width = 40
    ws_plan.column_dimensions['C'].width = 25
    ws_plan.column_dimensions['D'].width = 15
    ws_plan.column_dimensions['E'].width = 15
    ws_plan.column_dimensions['F'].width = 15
    ws_plan.column_dimensions['G'].width = 15
    for i in range(len(dates)):
        col_letter = get_column_letter(8 + i)
        ws_plan.column_dimensions[col_letter].width = 6

    # ---------------------------------------------------------
    # PREPARAR DATOS PARA INFORME Y DASHBOARD
    # ---------------------------------------------------------
    # Solo OP con Fecha Fin válida
    valid_ops = [op for op in opsData if op['details'].get('fecha_fin') and op['details']['fecha_fin'] != '9999-12-31']
    # Ordenar por fecha_fin
    valid_ops.sort(key=lambda x: x['details']['fecha_fin'])

    # ---------------------------------------------------------
    # HOJA 2: INFORME (MATRIZ FECHAS VS MAQUINAS)
    # ---------------------------------------------------------
    ws_inf = wb.create_sheet("INFORME")
    
    # Obtener máquinas únicas usadas
    used_machines = set()
    for op in opsData:
        for task in op['tasks']:
            m_name = task['machine']
            if re.match(r'^0+$', m_name):
                m_name = '0'
            used_machines.add(m_name)
            
    sorted_machines = sorted(list(used_machines))
    
    headers_inf = ["Maquina"] + sorted_machines
    ws_inf.append(headers_inf)
    
    for col_num in range(1, len(headers_inf) + 1):
        cell = ws_inf.cell(row=1, column=col_num)
        cell.fill = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws_inf.freeze_panes = "B2"
    
    matrix = {}
    for d in dates:
        matrix[d] = {}
        
    for op in opsData:
        op_id = op['details'].get('op', '')
        for task in op['tasks']:
            m_name = task['machine']
            if re.match(r'^0+$', m_name):
                m_name = '0'
            for date_str in task['all_dates']:
                if date_str in matrix:
                    if m_name in matrix[date_str]:
                        matrix[date_str][m_name] += f" / {op_id}"
                    else:
                        matrix[date_str][m_name] = op_id
                        
    row_idx_inf = 2
    for date_str in dates:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        row_data = [dt.strftime('%d/%m/%Y')]
        for mach in sorted_machines:
            row_data.append(matrix[date_str].get(mach, ""))
        ws_inf.append(row_data)
        row_idx_inf += 1
        
    ws_inf.column_dimensions['A'].width = 15
    for i in range(len(sorted_machines)):
        col_letter = get_column_letter(2 + i)
        ws_inf.column_dimensions[col_letter].width = 15
        
    if len(dates) > 0 and len(headers_inf) > 0:
        set_border(ws_inf, f"A1:{get_column_letter(len(headers_inf))}{row_idx_inf - 1}")

    # Calculos para DASHBOARD
    conteo_mensual = {}
    conteo_por_tipo = {}
    lista_detallada = []
    
    tipos_target = ["MAZA-XM", "MAZA-CONV", "CASCO-XM", "CASCO"]
    
    for op in valid_ops:
        d = op['details']
        ff_str = d['fecha_fin']
        ff_dt = datetime.strptime(ff_str, '%Y-%m-%d')
        
        mes_texto = get_month_name(ff_dt)
        anio = str(ff_dt.year)
        llave_mes = f"{anio} | {mes_texto}"
        tipo_raw = d.get('tipo', '').upper()
        
        # Conteo Total
        conteo_mensual[llave_mes] = conteo_mensual.get(llave_mes, 0) + 1
        
        # Conteo Tipo
        if llave_mes not in conteo_por_tipo:
            conteo_por_tipo[llave_mes] = {t: 0 for t in tipos_target}
        for t in tipos_target:
            if t in tipo_raw:
                conteo_por_tipo[llave_mes][t] += 1
                
        # Lista detallada DASHBOARD
        lista_detallada.append([
            d.get('op', ''),
            d.get('cliente', ''),
            d.get('eje', ''),
            tipo_raw,
            d.get('gr1_maquina', 'N/A'),
            ff_dt.strftime('%d/%m/%Y'),
            mes_texto
        ])

    # ---------------------------------------------------------
    # HOJA 3: DASHBOARD
    # ---------------------------------------------------------
    ws_dash = wb.create_sheet("DASHBOARD")
    ws_dash.sheet_view.showGridLines = False
    
    current_row = 2
    
    # --- TABLA 1: RESUMEN CANTIDADES TOTALES ---
    ws_dash.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    title_cell = ws_dash.cell(row=current_row, column=2, value="REPORTE DE ENTREGAS (CANTIDADES TOTALES POR MES)")
    title_cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    title_cell.font = Font(color="FFFFFF", bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    current_row += 1
    ws_dash.cell(row=current_row, column=2, value="Mes / Año").fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    ws_dash.cell(row=current_row, column=3, value="Total OPs").fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    ws_dash.cell(row=current_row, column=2).font = Font(bold=True)
    ws_dash.cell(row=current_row, column=3).font = Font(bold=True)
    
    start_border_t1 = current_row - 1
    
    meses_ordenados = sorted(conteo_mensual.keys())
    for mes in meses_ordenados:
        current_row += 1
        ws_dash.cell(row=current_row, column=2, value=mes)
        ws_dash.cell(row=current_row, column=3, value=conteo_mensual[mes])
        
    set_border(ws_dash, f"B{start_border_t1}:C{current_row}")
    current_row += 3
    
    # --- TABLA 2: RESUMEN POR TIPO ---
    start_border_t2 = current_row
    ws_dash.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
    title_cell2 = ws_dash.cell(row=current_row, column=2, value="RESUMEN DE PRODUCTOS POR TIPO Y MES")
    title_cell2.fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    title_cell2.font = Font(color="FFFFFF", bold=True)
    title_cell2.alignment = Alignment(horizontal="center", vertical="center")
    
    current_row += 1
    headers_t2 = ["Mes / Año", "MAZA-XM", "MAZA-CONV", "CASCO-XM", "CASCO"]
    for i, h in enumerate(headers_t2):
        c = ws_dash.cell(row=current_row, column=2+i, value=h)
        c.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        c.font = Font(bold=True)
        
    for mes in meses_ordenados:
        current_row += 1
        c_tipo = conteo_por_tipo[mes]
        ws_dash.cell(row=current_row, column=2, value=mes)
        ws_dash.cell(row=current_row, column=3, value=c_tipo.get("MAZA-XM", 0))
        ws_dash.cell(row=current_row, column=4, value=c_tipo.get("MAZA-CONV", 0))
        ws_dash.cell(row=current_row, column=5, value=c_tipo.get("CASCO-XM", 0))
        ws_dash.cell(row=current_row, column=6, value=c_tipo.get("CASCO", 0))
        
    set_border(ws_dash, f"B{start_border_t2}:F{current_row}")
    current_row += 3
    
    # --- TABLA 3: LISTADO DETALLADO ---
    start_border_t3 = current_row
    ws_dash.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=8)
    title_cell3 = ws_dash.cell(row=current_row, column=2, value="LISTADO DETALLADO DE OPs POR FECHA DE ENTREGA")
    title_cell3.fill = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid")
    title_cell3.font = Font(color="FFFFFF", bold=True)
    title_cell3.alignment = Alignment(horizontal="center", vertical="center")
    
    current_row += 1
    headers_t3 = ["OP", "Cliente", "Eje", "Tipo", "GR1 Maquina", "Fecha Fin", "Mes de Entrega"]
    for i, h in enumerate(headers_t3):
        c = ws_dash.cell(row=current_row, column=2+i, value=h)
        c.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        c.font = Font(bold=True)
        
    for item in lista_detallada:
        current_row += 1
        for i, val in enumerate(item):
            ws_dash.cell(row=current_row, column=2+i, value=val)
            
    set_border(ws_dash, f"B{start_border_t3}:H{current_row}")
    
    # Anchos de columna DASHBOARD
    ws_dash.column_dimensions['A'].width = 5
    ws_dash.column_dimensions['B'].width = 25
    ws_dash.column_dimensions['C'].width = 35
    ws_dash.column_dimensions['D'].width = 15
    ws_dash.column_dimensions['E'].width = 20
    ws_dash.column_dimensions['F'].width = 20
    ws_dash.column_dimensions['G'].width = 15
    ws_dash.column_dimensions['H'].width = 20

    # Guardar a bytes
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
